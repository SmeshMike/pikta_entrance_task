import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import logger as lg


def find_shape_of_data(data: dict) -> int:
    """
    Находит число заголовков по количеству элементов в header из переданного json.

    :param data: словарь полученный из json
    :returns: количество заголовков
    """
    x = len(data["headers"])
    return x


def align_data(data: dict) -> tuple[dict]:
    """
    Переформировывает переданный json файл, совмещая данные из заголовка и значений, сортируя по X и Y.

    Первые N элементов будут соответствовать заголовку, а все последующие комплекты из N элементов будут
    соответствовать последующим строкам значений.

    :param data: словарь полученный из json
    :returns: tuple из dict'ов всех значений из изначального json
    """
    all_elements = [properties["properties"] for elem in data for properties in data[elem]]
    return tuple(sorted(all_elements, key=lambda d: (int(d["Y"]), int(d["X"]))))


def fill_worksheet_header(length: int, data: tuple[dict], field: str, ws: Worksheet, bold=True) -> None:
    """
    Заполняет заголовочные поля рабочего листа.

    Записывает названия столбцов, формирует ширину столбцов по чилу элементов заголовка, увеличенного в 1.5 раза.

    :param length: число столбцов заголовка
    :param data: контент заголовков
    :param field: поле, по которому записываются названия
    :param ws: рабочий лист
    :param bold: делать ли шрифт жирным
    """
    for i in range(length):
        cell = ws.cell(row=1, column=i + 1, value=data[i][field])
        cell.font = Font(bold=bold)
        ws.column_dimensions[get_column_letter(i + 1)].width = int(len(str(data[i][field])) * 1.5)


def fill_worksheet_values(start: int, data: tuple[dict], field: str, ws: Worksheet) -> None:
    """
    Заполняет поля рабочего листа: записывает значения столбцов.

    :param start: элемент начала заполнения
    :param data: контент значений
    :param field: поле, по которому записываются названия
    :param ws: рабочий лист
    """
    for i in range(0, len(data)):
        value = data[i][field]
        if isinstance(value, str) and value[-1] == "-":
            value = "-" + value[: (len(value) - 1)]
        ws.cell(row=i // start + 2, column=i % start + 1, value=value)


def add_borders_to_cells(ws: Worksheet) -> None:
    """
    Добвляет в таблицу границы клеток.

    :param ws: рабочий лист
    """
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col in ws.iter_cols():
        for cell in col:
            cell.border = thin_border


def create_and_fill_sheet(data: dict, wb: Workbook, sheet_name: str) -> None:
    """
    Создаёт, заполняет и сохраняет xlsx файл по переданному словарю в директорию файла.

    :param data: словарь значений для записи
    :param wb: рабочая книга
    :param sheet_name: имя листа для сохранения
    """
    ws = wb.create_sheet(sheet_name)
    elements = align_data(data)
    x = find_shape_of_data(data)

    fill_worksheet_header(x, elements[:x], "QuickInfo", ws)
    fill_worksheet_values(x, elements[x:], "Text", ws)
    add_borders_to_cells(ws)


def convert_jsons_to_xlsx(files: list[str], name: str) -> None:
    """
    Основная функция запуска.

    Создаёт xlsx файл и по данным из json'ов в текущей директории.
    Каждый json заполняет новый лист.

    :param files: словарий значений
    """
    try:
        if files:
            wb = Workbook()
            logger = lg.get_logger()
            for file in files:
                with open(file, encoding="utf-8") as f:
                    data = json.load(f)
                    sheet_name = file.split(".")[0]
                    create_and_fill_sheet(data, wb, sheet_name)

            del wb["Sheet"]
            wb.save(f"{name}.xlsx")
            logger.info("Файл успешно сохранён")
    except KeyError:
        logger.info("Проверьте целостность файлов")


if __name__ == "__main__":
    files = [file for file in os.listdir(".") if file.endswith(".json")]
    convert_jsons_to_xlsx(files, "MyFile")
