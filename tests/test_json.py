import os
import sys
from unittest.mock import MagicMock, mock_open, patch

import pytest
from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(SCRIPT_DIR, ".."))

import json_task as jt


@pytest.mark.parametrize(
    "data, expected",
    [({"headers": ["1", "2"]}, 2), ({"headers": ["1", "2", "3", "4"]}, 4)],
)
def test_find_shape_of_data(data, expected):
    assert jt.find_shape_of_data(data) == expected


@pytest.mark.parametrize(
    "data",
    [
        ({"headers": [{"properties": {"Hotspot": "true", "VisualType": "text"}}]}),
        ({"headers": [{"smth": {"Hotspot": "true", "VisualType": "text"}}]}),
    ],
)
def test_align_data_for_exceptions(data):
    with pytest.raises(KeyError):
        jt.align_data(data)


@pytest.mark.parametrize(
    "data, expected",
    [
        ({"headers": [{"properties": {"X": 1, "Y": 2}}]}, 1),
        ({"headers": [{"properties": {"X": 1, "Y": 2}}, {"properties": {"X": 2, "Y": 3}}]}, 2),
    ],
)
def test_align_data(data, expected):
    assert len(jt.align_data(data)) == expected


@pytest.mark.parametrize(
    "data, expected",
    [
        (({"Text": 1}, {"Text": 2}), 2),
        (({"Text": 1}, {"Text": 2}, {"Text": 3}, {"Text": 4}), 4),
    ],
)
def test_fulfilling_worksheet(data, expected):
    wb = Workbook()
    ws = wb.create_sheet()
    jt.fill_worksheet_header(len(data), data, "Text", ws)
    jt.fill_worksheet_values(len(data), data, "Text", ws)
    assert ws.max_column == expected
    assert ws.max_row == 2


@pytest.mark.parametrize(
    "data",
    [
        (({"Text": 1}, {"Text": 2})),
        (({"Text": 1}, {"Text": 2}, {"Text": 3}, {"Text": 4})),
    ],
)
def test_add_borders_to_cells(data):
    wb = Workbook()
    ws = wb.create_sheet()
    jt.fill_worksheet_header(len(data), data, "Text", ws)
    jt.add_borders_to_cells(ws)
    for col in ws.iter_cols():
        for cell in col:
            assert cell.border.left.style is not None
